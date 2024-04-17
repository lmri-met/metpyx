import tkinter as tk
print("5%")
from tkinter import messagebox, filedialog, ttk
import serial
from tkinter import font as tkFont  # Importamos tkFont para medir el ancho del texto
import datetime
import statistics
print("25%")
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
import pandas as pd
print("50%")
import subprocess
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.backends.backend_pdf import PdfPages
import os
print("75%")
import win32com.client
import random
import threading
print("100%")

class ElectrometroGUI:
    def __init__(self, master):
        try:
            os.remove('datos.txt')
            print("Archivo 'datos.txt' borrado exitosamente al iniciar.")
        except FileNotFoundError:
            print("Archivo 'datos.txt' no encontrado; iniciando nuevo archivo.")
        except OSError as e:
            print(f"Error al borrar 'datos.txt' al iniciar: {e}")
    
        self.master = master
        self.master.title("Cámara IG-11 - Medida de Series Temporales - Emisores Gamma")
        self.table = None

        # Create a new frame for the title
        title_frame = tk.Frame(self.master)
        title_frame.grid(row=0, column=0, columnspan=5)  # Changed pack to grid

        # Create a label for the title
        title_label = tk.Label(title_frame, text="Cámara IG-11\nMedida de Series Temporales\nEmisores Gamma", font=("Helvetica", 16, "bold"))
        title_label.pack()

        # Create a new frame for the other widgets
        self.widget_frame = tk.Frame(self.master)
        self.widget_frame.grid(row=1, column=0, columnspan=5)  # Changed pack to grid
        
        # self.serial_port = 'COM8'  # Reemplaza con el puerto correcto
        # self.ser = serial.Serial(self.serial_port, 9600, timeout=1)

        self.medida_nombre_var = tk.StringVar(value='')  # Valor predeterminado
        self.medida_persona_var = tk.StringVar(value='')  # Valor predeterminado
        self.medida_fecha_var = tk.StringVar(value='')  # Valor predeterminado
        self.tipo_medida_var = tk.StringVar(value='Fondo')  # Valor predeterminado
        self.num_serie_var = tk.StringVar(value='1')  # Valor predeterminado
        self.tiempo_entre_medidas_var = tk.StringVar(value='1')  # Valor predeterminado
        self.num_medidas_var = tk.StringVar(value='10')  # Valor predeterminado
        self.todos_valores_medidos = []
        self.valores_medios_por_serie = []
        self.ultimo_tiempo_medido = None
        self.num_medidas_realizadas = 0
        self.valores_medidos = []
        self.valor_medio = 0
        self.sigma_medio = 0
        self.num_serie_actual = 1
        self.start_time = None
        self.running = False  # Variable para controlar si se está tomando datos

        self.color_index = 0  # Índice para el color actual de la gráfica
        self.colores = ['blue', 'green', 'red', 'cyan', 'magenta', 'yellow', 'black']  # Lista de colores para las gráficas
        self.tandas = []

        self.button_container = None  # Inicializamos el atributo button_container

        self.archivo_datos = None  # Inicializa como None
        self.nombre_archivo_datos = f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_medidas.txt"  # Nombre de archivo basado en la hora de inicio

        self.create_widgets()

    def create_widgets(self):
        # Campo para introducir el nombre de la medida
        tk.Label(self.widget_frame, text="Nombre de la Medida:").grid(row=0, column=0)
        self.nombre_medida_entry = tk.Entry(self.widget_frame, textvariable=self.medida_nombre_var)
        self.nombre_medida_entry.grid(row=0, column=1)

        # Campo para introducir la persona que realiza la medida
        tk.Label(self.widget_frame, text="Persona que realiza la medida:").grid(row=1, column=0)
        self.persona_medida_entry = tk.Entry(self.widget_frame, textvariable=self.medida_persona_var)
        self.persona_medida_entry.grid(row=1, column=1)

        # Campo para mostrar la fecha y hora de la medida
        tk.Label(self.widget_frame, text="Fecha y Hora de la medida:").grid(row=2, column=0)
        self.fecha_hora_medida_label = tk.Label(self.widget_frame, textvariable=self.medida_fecha_var)
        self.fecha_hora_medida_label.grid(row=2, column=1)

        # Campo para seleccionar el puerto COM
        tk.Label(self.widget_frame, text="Puerto COM:").grid(row=3, column=0)  # Nota el cambio de índice de fila a 3
        opciones_puerto_com = ["SIMULACIÓN"] + [f'COM{i}' for i in range(1, 11)]
        self.puerto_com_var = tk.StringVar(value="SIMULACIÓN")  # Cambiado a "SIMULACIÓN" como valor predeterminado
        self.puerto_com_optionmenu = tk.OptionMenu(self.widget_frame, self.puerto_com_var, *opciones_puerto_com)
        self.puerto_com_optionmenu.grid(row=3, column=1)

        # Campo para seleccionar el puerto COM
        # tk.Label(self.widget_frame, text="Puerto COM:").grid(row=3, column=0)  # Nota el cambio de índice de fila a 3
        # self.puerto_com_var = tk.StringVar(value='COM5')  # Valor predeterminado
        # self.puerto_com_optionmenu = tk.OptionMenu(self.widget_frame, self.puerto_com_var, *[f'COM{i}' for i in range(1, 11)])
        # self.puerto_com_optionmenu.grid(row=3, column=1)

        # Campo para elegir el tipo de medida
        tk.Label(self.widget_frame, text="Tipo de medida:").grid(row=4, column=0)
        self.tipo_medida_optionmenu = tk.OptionMenu(self.widget_frame, self.tipo_medida_var, "Fondo", "Nucleido")
        self.tipo_medida_optionmenu.grid(row=4, column=1)

        # Campo para introducir el número de serie de medidas
        tk.Label(self.widget_frame, text="Número de serie de medidas:").grid(row=5, column=0)
        self.num_serie_entry = tk.Entry(self.widget_frame, textvariable=self.num_serie_var)
        self.num_serie_entry.grid(row=5, column=1)

        # Campo para introducir el tiempo entre series
        tk.Label(self.widget_frame, text="Tiempo entre series (segundos):").grid(row=6, column=0)
        self.tiempo_entre_medidas_entry = tk.Entry(self.widget_frame, textvariable=self.tiempo_entre_medidas_var)
        self.tiempo_entre_medidas_entry.grid(row=6, column=1)

        # Campo para introducir el número de medidas
        tk.Label(self.widget_frame, text="Número de medidas:").grid(row=7, column=0)
        self.num_medidas_entry = tk.Entry(self.widget_frame, textvariable=self.num_medidas_var)
        self.num_medidas_entry.grid(row=7, column=1)

        # Botón para iniciar la adquisición de datos
        self.start_button = tk.Button(self.master, text="Iniciar", command=self.start_acquisition, bg='lightgreen', font=("Helvetica", 12, "bold"))
        self.start_button.grid(row=11, column=5, pady=5, padx=10)

        # Botón para detener la adquisición de datos
        self.stop_button = tk.Button(self.master, text="Detener", command=self.stop_acquisition, bg='red', font=("Helvetica", 12, "bold"))
        self.stop_button.grid(row=11, column=5, pady=5, padx=10)
        self.stop_button.grid_remove()  # Hacer el botón invisible inicialmente

        # Crear un contenedor para los botones
        self.button_container = tk.Frame(self.master)

        # Botón para reiniciar la medida
        self.restart_button = tk.Button(self.button_container, text="Reiniciar medida", command=self.restart_program, bg='orange',font=("Helvetica", 12, "bold"))
        self.restart_button.grid(row=0, column=0)  # Utilizamos grid en lugar de pack

        # Botón para crear fichero Excel
        self.save_excel_button = tk.Button(self.button_container, text="Guardar Excel", command=self.save_to_excel, state='disabled', bg='white',font=("Helvetica", 12, "bold"))
        self.save_excel_button.grid(row=1, column=0)  # Utilizamos grid en lugar de pack

        # Colocar el contenedor en la misma fila y columna
        self.button_container.grid(row=11, column=5)

        # Ocultar los botones
        self.button_container.grid_remove()

        # Etiqueta para mostrar los resultados
        self.result_label = tk.Label(self.master, text="")
        self.result_label.grid(row=8, columnspan=4)

        # Etiqueta para mostrar los datos en directo
        self.live_data_label = tk.Label(self.master, text="", font=("Helvetica", 12, "bold"))
        self.live_data_label.grid(row=10, columnspan=4)

        # Gráfico para visualizar los datos
        self.fig = plt.figure(figsize=(7, 4))
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.master)
        self.canvas.get_tk_widget().grid(row=0, column=5, rowspan=10, padx=20)

        self.create_table()

    def start_acquisition(self):
        print("Inicio de la adquisición de datos...")
        self.abrir_archivo_datos()  # Abrir el archivo de datos
        # Obtener la fecha y hora actual
        fecha_hora_actual = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.medida_fecha_var.set(fecha_hora_actual)

        # Configurar el inicio del tiempo
        # Inicializar los tiempos de inicio y fin de la primera y última serie
        self.start_time_first_serie = datetime.datetime.now()
        self.end_time_last_serie = self.start_time_first_serie
        self.start_time = time.time()
        self.running = True

        # Hide the start button and show the stop button
        self.start_button.grid_remove()
        self.stop_button.grid()

        # Realizar la adquisición de datos
        num_serie = int(self.num_serie_var.get())
        tiempo_entre_medidas = int(self.tiempo_entre_medidas_var.get())
        num_medidas = int(self.num_medidas_var.get())
        self.acquire_data(num_serie, tiempo_entre_medidas, num_medidas)
        

        # Al iniciar una nueva adquisición, guarda la tanda actual y resetea para la próxima
        if self.todos_valores_medidos:
            self.tandas.append({
                'datos': self.todos_valores_medidos.copy(),
                'color': self.colores[self.color_index % len(self.colores)],
                'tipo_medida': self.tipo_medida_var.get()
            })
            self.color_index += 1  # Prepara el color para la próxima tanda
        self.todos_valores_medidos.clear()  # Prepara los datos para la próxima tanda


    def stop_acquisition(self):
        # Detener la adquisición de datos
        self.running = False

        # Ocultar el botón de detener y mostrar el botón de iniciar (si fuera necesario)
        self.stop_button.grid_remove()
        self.start_button.grid()

        self.cerrar_archivo_datos()  # Cerrar el archivo de datos

        # Insertar los datos de la última serie en la tabla (si aplica)
        self.insert_last_series_data()

        # Reiniciar todas las variables y actualizar la GUI (si aplica)
        self.num_medidas_realizadas = 0
        self.valores_medidos = []
        self.valor_medio = 0
        self.sigma_medio = 0
        self.num_serie_actual = 1
        self.start_time = None
        self.result_label.config(text="")
        self.live_data_label.config(text="")
        self.ax.clear()
        self.canvas.draw()



        # Mostrar una ventana de advertencia o información con el mensaje final
        messagebox.showinfo("Fin de la adquisición", "La toma de datos ha terminado.", master=self.master)

        # Cerrar la aplicación después de mostrar el mensaje
        self.master.destroy()

    def abrir_archivo_datos(self):
        # Generar nombre del archivo basado en el nombre de la medida y la fecha/hora actual
        self.nombre_archivo_datos = 'datos.txt'
        if self.archivo_datos is None:  # Solo abre si no está ya abierto
                try:
                    # Abre el archivo en modo de añadir para no sobrescribir los datos existentes
                    self.archivo_datos = open(self.nombre_archivo_datos, 'a')  
                    if self.archivo_datos.tell() == 0:  # Si el archivo está vacío, añade el encabezado
                        encabezado = "Nº Medida   Serie       Media       Sigma       Tiempo      Tipo de medida\n"
                        self.archivo_datos.write(encabezado)
                except Exception as e:
                    print(f"Error al abrir el archivo: {e}")
                    self.archivo_datos = None

    def escribir_datos(self, serie, numero_serie, medida, sigma, tiempo, tipo):
        if self.archivo_datos is not None:
            try:
                # Formatear datos con 5 decimales y coma como separador decimal
                medida_str = f"{medida:.5e}".replace('.', ',').ljust(12)[:12]
                sigma_str = f"{sigma:.5e}".replace('.', ',').ljust(12)[:12]
                # Formatear el resto de los datos para asegurar 12 caracteres
                serie_str = str(serie).ljust(12)[:12]
                numero_serie_str = str(numero_serie).ljust(12)[:12]
                tiempo_str = str(tiempo).ljust(12)[:12]
                tipo_str = str(tipo).ljust(12)[:12]
                
                # Escribir en el archivo
                self.archivo_datos.write(f"{serie_str}{numero_serie_str}{medida_str}{sigma_str}{tiempo_str}{tipo_str}\n")
                self.archivo_datos.flush()  # Asegurarse de que se escribe en el archivo inmediatamente
            except Exception as e:
                print(f"Error al escribir en el archivo: {e}")

    def cerrar_archivo_datos(self):
        if self.archivo_datos is not None:
            self.archivo_datos.close()
            self.archivo_datos = None

    def insert_last_series_data(self):
        # Obtener la fecha y hora actual
        fecha_hora_actual = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Obtener la duración de la última serie
        end_time_serie = datetime.datetime.now()
        start_time_serie = end_time_serie - datetime.timedelta(seconds=int(self.num_medidas_var.get()) * int(self.tiempo_entre_medidas_var.get()))
        time_diff = end_time_serie - start_time_serie
        time_diff_seconds = time_diff.total_seconds()

        # Insertar los datos en la tabla
        self.table.insert("", "end", values=(self.num_serie_actual, f"{self.valor_medio:.4e}", f"{self.sigma_medio:.4e}",
                                        start_time_serie.strftime("%H:%M:%S"), end_time_serie.strftime("%H:%M:%S"),
                                        f"{time_diff_seconds:.1f}", self.tipo_medida_var.get()))  # Se añade el tipo de medida
        
        
        pass

    def restart_program(self):
        # print("Restarting program...")  # Mensaje de depuración
    
        # Hide the restart button
        self.restart_button.grid_remove()

        # Hide the save button
        self.save_excel_button.grid_remove()

        # Ocultar el contenedor de botones
        self.button_container.grid_remove()

        # Show the start button
        self.start_button.grid()

        # Reset all the variables
        self.num_medidas_realizadas = 0
        self.valores_medidos = []
        self.valor_medio = 0
        self.sigma_medio = 0
        self.num_serie_actual = 1
        self.start_time = None

        # Incrementa el índice del color para cambiar el color en la próxima serie de medidas
        self.color_index += 1
        
        # Reset the GUI
        self.result_label.config(text="")
        self.live_data_label.config(text="")
        self.ax.clear()
        self.canvas.draw()

        # print("Showing start button...")  # Mensaje de depuración
        # Show the button container only if it's currently visible
        if self.button_container.winfo_viewable():
            self.button_container.grid() 

    def create_table(self):
        # Create a new frame for the table
        table_frame = tk.Frame(self.master)
        table_frame.grid(row=11, column=0, columnspan=5)  # Cambiado el número de columnas a 3

        # Create the table without the "Tipo de medida" column
        self.table = ttk.Treeview(table_frame, columns=("Serie", "Media", "Sigma", "Inicio", "Final", "Duración", "Tipo de medida"), show="headings")
        self.table.heading("Tipo de medida", text="Tipo de medida")  # Agregar el encabezado de la columna
        self.table.column("Tipo de medida", width=100)  # Ajustar el ancho de la columna
            
        # Set the column headings
        self.table.heading("Serie", text="Serie")
        self.table.heading("Media", text="Media")
        self.table.heading("Sigma", text="Sigma")
        self.table.heading("Inicio", text="Inicio")
        self.table.heading("Final", text="Final")
        self.table.heading("Duración", text="Duración")

        # Pack the table
        self.table.pack(fill="both", expand=True)

        # Initialize dictionary to store maximum column widths
        column_widths = {column: len(column.title()) for column in ("Serie", "Media", "Sigma", "Inicio", "Final", "Duración")}

        # Update column widths with the maximum width of the data
        for row in self.table.get_children():
            for column, value in zip(("Serie", "Media", "Sigma", "Inicio", "Final", "Duración"), self.table.item(row)['values']):
                column_widths[column] = max(column_widths[column], len(str(value)))

        # Center the contents of each column
        for column in ("Serie", "Media", "Sigma", "Inicio", "Final", "Duración"):
            self.table.column(column, anchor="center")

        # Adjust column widths
        for column in column_widths:
            width = int(tkFont.Font().measure(column.title()) * 1.4) + 10  # Add extra padding and increase by 10%
            self.table.column(column, width=width)

    def acquire_data(self, num_serie, tiempo_entre_medidas, num_medidas):
        total_measures = num_serie * num_medidas
        print(f"Iniciando la adquisición de la serie {num_serie}...")
        self.valores_medios_tanda_actual = []
        self.sigmas_medios_tanda_actual = []
        # Limpiar la gráfica al iniciar una nueva adquisición
        self.ax.clear()
        self.ax.set_xlabel("Número de Medida")
        self.ax.set_ylabel("Amperaje")

        for serie in range(1, num_serie + 1):
            print(f"Procesando serie {serie}...")
            self.valores_medidos_serie_actual = []
            self.num_serie_actual = serie  # Actualizar el número de serie actual
            # Reiniciar la lista de valores medidos al inicio de cada serie
            self.valores_medidos = []
            self.valores_medios_tanda_actual = []  # Resetear los valores medios para la nueva tanda
            self.sigmas_medios_tanda_actual = []  # Resetear los sigmas medios para la nueva tanda

            
            self.end_time_last_serie = datetime.datetime.now()
            start_time_serie = datetime.datetime.now()  # Start time for the series
            for medida in range(1, num_medidas + 1):
                if not self.running:  # Verificar si se debe detener
                    return  # Salir de la función si se detiene la adquisición
                
                seleccion = self.puerto_com_var.get()  # Obtener la selección actual del menú desplegable
                if seleccion == "SIMULACIÓN":
                    # Simular la lectura de datos del electrometro con números aleatorios entre 1 y 10
                    valor_medido = random.randint(1, 10)  # Generar un número aleatorio

                else:
                    # Asumiendo que self.ser es el objeto serial para la comunicación con el puerto COM
                    self.ser.write(b"*RST\n")  # Resetear el electrometro a su estado predeterminado
                    self.ser.write(b":SENS:FUNC 'CURR:DC'\n")  # Establecer la función de medición a corriente DC
                    self.ser.write(b":SENS:CURR:RANG 1e-6\n")  # Establecer el rango de medición a 1 microamperio
                    self.ser.write(b":READ?\n")  # Realizar la lectura
                    lectura = self.ser.readline().strip().decode()  # Leer la respuesta
                    valor_medido = float(lectura.split(',')[0])  # Extraer el valor medido
                
                # Simular la lectura de datos del electrometro con números aleatorios entre 1 y 10
                #valor_medido = random.randint(1, 10)  # Generar un número aleatorio

                # Las siguientes líneas quedan comentadas porque ahora se simula la adquisición de datos
                # self.ser.write(b"*RST\n")  # Resetear el electrometro a su estado predeterminado
                # self.ser.write(b":SENS:FUNC 'CURR:DC'\n")  # Establecer la función de medición a corriente DC
                # self.ser.write(b":SENS:CURR:RANG 1e-6\n")  # Establecer el rango de medición a 1 microamperio
                # self.ser.write(b":READ?\n")  # Realizar la lectura
                # lectura = self.ser.readline().strip().decode()  # Leer la respuesta
                # valor_medido = float(lectura.split(',')[0])  # Extraer el valor medido

                self.todos_valores_medidos.append(valor_medido)
                self.valores_medios_por_serie.append(self.valor_medio)
                self.valores_medidos.append(valor_medido)
                print(f"Valor medido: {valor_medido}")
                self.num_medidas_realizadas += 1

                # Calcular el valor medio
                self.valor_medio = statistics.mean(self.valores_medidos)

                # Calcular el Sigma medio solo si hay suficientes datos
                if len(self.valores_medidos) >= 2:
                    self.sigma_medio = statistics.stdev(self.valores_medidos) / (len(self.valores_medidos) ** 0.5)
                    sigma = self.sigma_medio

                # Actualizar la etiqueta de resultados
                self.result_label.config(text=self.format_results())

                # Actualizar la etiqueta de datos en directo
                self.live_data_label.config(text=f"Última medida directa: {valor_medido}")

                # Actualizar la gráfica
                self.update_graph()

                # Habilitar el botón de guardar si se han realizado todas las medidas
                if self.num_medidas_realizadas >= total_measures:
                    self.stop_button.grid_remove()
                    self.save_excel_button.grid()
                    self.save_excel_button.config(state='normal')  # Enable the save button
                    self.restart_button.grid()  # Show the restart button

                # Actualizar la GUI
                self.master.update()

                # Esperar el tiempo entre medidas
                time.sleep(tiempo_entre_medidas)
                self.ultimo_tiempo_medido = datetime.datetime.now()
                tiempo= self.ultimo_tiempo_medido.strftime("%H:%M:%S")
                self.escribir_datos(medida, serie, valor_medido, sigma=self.sigma_medio, tiempo=tiempo, tipo=self.tipo_medida_var.get())

            print(f"Valor medio de la serie {serie}: {self.valor_medio}")
            end_time_serie = datetime.datetime.now()  # End time for the series
            time_diff = end_time_serie - start_time_serie  # Calculate time difference
            time_diff_seconds = time_diff.total_seconds()  # Convert time difference to seconds
            valor_medio_serie_actual = statistics.mean(self.valores_medidos)
            sigma_medio_serie_actual = statistics.stdev(self.valores_medidos) if len(self.valores_medidos) > 1 else 0

            # Ahora que ya están definidas, añádelas a las listas de la tanda actual
            self.valores_medios_tanda_actual.append(valor_medio_serie_actual)
            self.sigmas_medios_tanda_actual.append(sigma_medio_serie_actual)
            # Add a new row to the table with time difference
            self.table.insert("", "end", values=(self.num_serie_actual, f"{self.valor_medio:.4e}", f"{self.sigma_medio:.4e}",
                                        start_time_serie.strftime("%H:%M:%S"), end_time_serie.strftime("%H:%M:%S"),
                                        f"{time_diff_seconds:.1f}", self.tipo_medida_var.get()))  # Se añade el tipo de medida
        # Mostrar el mensaje de finalización después de completar la tabla
        self.show_warning_message()

    def show_warning_message(self):
        print("Calculando el valor medio global y sigma medio global para la tanda actual...")

        valores_medios = []
        sigmas_medios = []
        start_time_first_serie = None
        ultimo_tiempo_medido = None

        # Iterar sobre las filas de la tabla para recopilar valores desde el inicio de la última tanda
        for child in self.table.get_children():
            serie, valor_medio, sigma_medio, inicio, final, *_ = self.table.item(child)['values']
            
            # Ignorar las filas "TOTAL"
            if serie == "TOTAL":
                valores_medios = []  # Reiniciar para la siguiente tanda
                sigmas_medios = []   # Reiniciar para la siguiente tanda
                continue
            
            # Convertir los valores medio y sigma a float y agregarlos a las listas
            valores_medios.append(float(valor_medio))
            sigmas_medios.append(float(sigma_medio))
            
            # Actualizar los tiempos de inicio y final de la tanda actual
            if start_time_first_serie is None:
                start_time_first_serie = inicio  # Tiempo de inicio de la primera serie en la tanda
            ultimo_tiempo_medido = final  # Tiempo de final de la última serie en la tanda

        if valores_medios and sigmas_medios:
            # Calcular el valor medio y sigma medio global de la tanda
            valor_medio_global_tanda = sum(valores_medios) / len(valores_medios)
            sigma_media_global_tanda = sum(sigmas_medios) / len(sigmas_medios)
            print
            
            # Calcular la duración total de la tanda
            if start_time_first_serie and ultimo_tiempo_medido:
                formato_tiempo = "%H:%M:%S"  # Asumiendo que los tiempos están en este formato
                tiempo_inicio = datetime.datetime.strptime(start_time_first_serie, formato_tiempo)
                tiempo_final = datetime.datetime.strptime(ultimo_tiempo_medido, formato_tiempo)
                duracion_total_seconds = (tiempo_final - tiempo_inicio).total_seconds()

            # Insertar la fila TOTAL en la tabla
            self.table.insert("", "end", values=(
                "TOTAL",
                f"{valor_medio_global_tanda:.4e}",
                f"{sigma_media_global_tanda:.4e}",
                start_time_first_serie,
                ultimo_tiempo_medido,
                f"{duracion_total_seconds:.1f}",
                self.tipo_medida_var.get(),  
                ""
            ))

            # Muestra mensaje de finalización y actualiza la UI
            messagebox.showinfo("Fin de la Toma de Datos", "La toma de datos de la tanda ha finalizado.")
            self.restart_button.grid()
            self.save_excel_button.grid()
            self.save_excel_button["state"] = "normal"
            self.button_container.grid()
        else:
            print("No hay datos suficientes para calcular el total de la tanda actual.")


    def update_graph(self):
        self.ax.clear()  # Limpia la gráfica para redibujar

        # Configura los títulos y etiquetas de los ejes
        self.ax.set_title("Mediciones")
        self.ax.set_xlabel("Número de Medida")
        self.ax.set_ylabel("Valor Medido")

        # Dibujar tandas anteriores con sus respectivos colores y etiquetas
        for tanda in self.tandas:
            datos = range(1, len(tanda['datos']) + 1)
            self.ax.plot(datos, tanda['datos'], marker='o', linestyle='-', color=tanda['color'], label=f'{tanda["tipo_medida"]}')
            # Agregar etiquetas en cada punto para tandas anteriores (opcional)
            for i, valor in enumerate(tanda['datos']):
                self.ax.text(i + 1, valor, f"{i + 1}", ha='center', va='bottom', fontsize=8, color=tanda['color'])

        # Dibujar la tanda actual si existe
        if self.todos_valores_medidos:
            datos_actuales = range(1, len(self.todos_valores_medidos) + 1)
            color_actual = self.colores[self.color_index % len(self.colores)]
            self.ax.plot(datos_actuales, self.todos_valores_medidos, marker='o', linestyle='-', color=color_actual, label=f'{self.tipo_medida_var.get()}')
            
            # Agregar etiquetas en cada punto para la tanda actual
            for i, valor in enumerate(self.todos_valores_medidos):
                etiqueta = f"M {i+1}/{len(self.todos_valores_medidos)}\nS {self.num_serie_actual}/{self.num_serie_var.get()}"
                self.ax.text(i + 1, valor, etiqueta, ha='center', va='bottom', fontsize=8, color=color_actual)

        self.ax.legend()  # Actualiza la leyenda con los tipos de medida
        self.canvas.draw()  # Redibuja el canvas

    def format_results(self):
        # Calcular tiempo transcurrido
        elapsed_time = time.time() - self.start_time
        # Calcular tiempo restante estimado (asumiendo un tiempo constante entre medidas)
        total_time = elapsed_time / self.num_medidas_realizadas * (int(self.num_medidas_var.get()) * int(self.num_serie_var.get()) - self.num_medidas_realizadas)
        # Convertir a segundos y porcentaje
        remaining_seconds = int(total_time)
        remaining_percentage = total_time / (elapsed_time + total_time) * 100 if elapsed_time + total_time > 0 else 0

        # Crear el texto de resultados
        result_str = "RESULTADOS\n"
        result_str += f"Valor medio: {self.valor_medio:.2e}   Sigma medio: {self.sigma_medio:.2e}\n"
        result_str += f"Serie actual: {self.num_serie_actual}\n"
        result_str += f"Medida actual: {self.num_medidas_realizadas}/{int(self.num_medidas_var.get()) * int(self.num_serie_var.get())}\n"
        result_str += f"Medidas realizadas: {self.num_medidas_realizadas}/{int(self.num_medidas_var.get()) * int(self.num_serie_var.get())}\n"
        result_str += f"Tiempo transcurrido: {elapsed_time:.2f} segundos\n"
        result_str += f"Tiempo restante: {remaining_seconds} segundos ({remaining_percentage:.2f}% del total)"
        return result_str
        
    def create_excel(self):
        # Generate the default filename
        default_filename = f"{self.medida_nombre_var.get()}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # Ask the user where to save the Excel file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_filename, filetypes=[("Excel files", "*.xlsx")])

        if not file_path:
            # The user cancelled the dialog
            return None

        return file_path

    def save_to_excel(self):
        # Asegúrate de cerrar el archivo de datos antes de intentar leerlo
        self.cerrar_archivo_datos()

        file_path = self.create_excel()
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        negrita_blanco_sobre_negro = Font(bold=True, color="FFFFFF")
        blanco_sobre_negro = Font(color="FFFFFF")
        fondo_negro = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Escribir los datos de entrada
        ws['A1'] = "Nombre de la Medida:"
        ws['B1'] = self.medida_nombre_var.get()
        ws['A2'] = "Persona que realiza la medida:"
        ws['B2'] = self.medida_persona_var.get()
        ws['A3'] = "Fecha y Hora de la medida:"
        ws['B3'] = self.medida_fecha_var.get()

        ws['A5'] = "Número de serie de medidas:"
        ws['B5'] = self.num_serie_var.get()
        ws['A6'] = "Tiempo entre series (segundos):"
        ws['B6'] = self.tiempo_entre_medidas_var.get()
        ws['A7'] = "Número de medidas:"
        ws['B7'] = self.num_medidas_var.get()
        ws['A8'] = "Equipos utilizados:"
        ws['B8'] = "Electrómetro Keithley 6517B (IR13-006)"
        ws['A9'] = "Procedimientos"
        ws['B9'] = "PLMRI-C-09 y PLMRI-C-10"

        for col in ['A', 'B']:
            max_length = 0
            column = ws[col]
            for cell in column:
                try: 
                    # Ajustar por el contenido de la celda
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Ajuste para que no quede demasiado justo
            ws.column_dimensions[col].width = adjusted_width

        # Asegurarse de que B5, B6 y B7 son numéricos reales
        ws['B5'].value = float(self.num_serie_var.get())  # Número de serie de medidas
        ws['B6'].value = float(self.tiempo_entre_medidas_var.get())  # Tiempo entre series (segundos)
        ws['B7'].value = float(self.num_medidas_var.get())  # Número de medidas

        # Aplicar estilos a todas las celdas pertinentes
        for row in range(1, 10):  # Asumiendo 9 filas de información basada en tu entrada
            ws[f'A{row}'].font = negrita_blanco_sobre_negro
            ws[f'A{row}'].fill = fondo_negro
            ws[f'B{row}'].font = blanco_sobre_negro
            ws[f'B{row}'].fill = fondo_negro

        # Estilo de fuente en negrita
        fuente_negrita = Font(bold=True)

        # Estilo de fondo verde claro
        fondo_verde_claro = PatternFill(start_color="E2EFDA",
                                        end_color="E2EFDA",
                                        fill_type="solid")

        # Escribir "Valor medio:" y "Sigma medio:" con formato
        ws['A10'] = "Valor medio:"
        ws['A10'].font = fuente_negrita
        ws['A10'].fill = fondo_verde_claro

        ws['A11'] = "Sigma medio:"
        ws['A11'].font = fuente_negrita
        ws['A11'].fill = fondo_verde_claro

        # Establecer los valores numéricos directamente para mantenerlos como números
        ws['B10'].value = self.valor_medio
        ws['B10'].font = fuente_negrita
        ws['B10'].fill = fondo_verde_claro
        # Aplicar formato de número con 5 decimales
        ws['B10'].number_format = '0.00000E+00'

        ws['B11'].value = self.sigma_medio
        ws['B11'].font = fuente_negrita
        ws['B11'].fill = fondo_verde_claro
        # Aplicar formato de número con 5 decimales
        ws['B11'].number_format = '0.00000E+00'

        # Definir los estilos
        fuente_negrita = Font(bold=True)
        fondo_naranja_claro = PatternFill(start_color="FFEB9C",
                                        end_color="FFEB9C",
                                        fill_type="solid")
        fondo_azul_claro = PatternFill(start_color="B6D7A8",
                               end_color="B6D7A8",
                               fill_type="solid")
        alineacion_centro = Alignment(horizontal="center")

        # Definir los encabezados de la tabla aquí para asegurar su disponibilidad
        table_columns = ("Serie", "Media", "Sigma", "Inicio", "Final", "Duración", "Tipo de medida")

        # Añadir el encabezado especial "VALORES MEDIOS" en la fila 14
        encabezado_especial = ws['A14']
        encabezado_especial.value = "VALORES MEDIOS"
        encabezado_especial.font = fuente_negrita
        encabezado_especial.fill = fondo_naranja_claro

        # Fusionar las celdas para el encabezado especial
        ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=len(table_columns))

        # Ahora, aplicar la alineación centrada a todas las celdas en el rango fusionado
        for row in ws.iter_rows(min_row=14, max_row=14, min_col=1, max_col=len(table_columns)):
            for cell in row:
                cell.alignment = alineacion_centro

        # Aplicar los estilos a los encabezados de la tabla, empezando en la fila 15
        for col_idx, col_name in enumerate(table_columns, start=1):
            celda = ws.cell(row=15, column=col_idx, value=col_name)
            celda.font = fuente_negrita
            celda.fill = fondo_naranja_claro

        # Asegurarte de que 'self.table.get_children()' y 'self.table.item(item)['values']' son accesibles
        # y representan la estructura de datos que deseas usar aquí
        for row_idx, item in enumerate(self.table.get_children(), start=16):
            for col_idx, val in enumerate(self.table.item(item)['values'], start=1):
                celda = ws.cell(row=row_idx, column=col_idx)

                # Aplicar el fondo azul claro a todas las celdas
                celda.fill = fondo_naranja_claro

                if isinstance(val, float):
                    # Trata el valor como numérico para Excel, pero formatea como string solo para visualización
                    if col_idx in (2, 3):  # Para 'Media' y 'Sigma' en formato exponencial
                        # Excel puede manejar esto como un número; solo estamos cambiando la representación visual aquí
                        formatted_value = f"{val:.5e}".replace('.', ',')
                        celda.value = formatted_value
                    elif col_idx == 6:  # Para 'Duración' con un solo decimal
                        # Igual que antes, Excel ve un número, pero cambiamos cómo se muestra
                        formatted_value = f"{val:.1f}".replace('.', ',')
                        celda.value = formatted_value
                else:
                    # Aquí asumimos que todos los otros valores deben ser tratados como numéricos también
                    # Intentamos convertir a un número, si es posible
                    try:
                        numeric_val = float(val)
                        celda.value = numeric_val
                    except ValueError:
                        # Si no es convertible a número, cae aquí
                        celda.value = val

        # Leer los datos del archivo 'datos.txt'
        # Cierra el archivo de datos antes de intentar borrarlo
        self.cerrar_archivo_datos()
        # Definir los encabezados y escribirlos en la fila 21, comenzando desde la columna I (9)
        encabezados = ["Media", "Serie", "Valor medida", "Sigma", "Tiempo", "Tipo de medida"]
        for col_index, encabezado in enumerate(encabezados, start=9):  
            celda = ws.cell(row=21, column=col_index)
            celda.value = encabezado
            celda.font = fuente_negrita
            celda.fill = fondo_azul_claro
            # Ajustar el ancho de la columna a 15
            ws.column_dimensions[celda.column_letter].width = 15

        # Leer los datos del archivo 'datos.txt'
        # Cierra el archivo de datos antes de intentar borrarlo
        self.cerrar_archivo_datos()
        # Definir los encabezados y escribirlos en la fila 21, comenzando desde la columna I (9)
        encabezados = ["Media", "Serie", "Valor medida", "Sigma", "Tiempo", "Tipo de medida"]
        for col_index, encabezado in enumerate(encabezados, start=9):  
            celda = ws.cell(row=21, column=col_index)
            celda.value = encabezado
            # Ajustar el ancho de la columna a 15
            ws.column_dimensions[celda.column_letter].width = 15

        # Leer los datos del archivo 'datos.txt'
        try:
            with open('datos.txt', 'r') as f:
                lines = f.readlines()
                # Comienza a escribir desde la fila 22, columna K (13), justo debajo de los encabezados
                for index, line in enumerate(lines[1:], start=22):  # Ignora el encabezado en el archivo de datos
                    for col_index, value in enumerate(line.strip().split(), start=9):  # Asumiendo espacio como delimitador
                        ws.cell(row=index, column=col_index).value = value
        except Exception as e:
            print(f"Error al leer 'datos.txt': {e}")

        # Guardar la gráfica como una imagen
        self.fig.savefig("graph.png")

        # Insertar la imagen en la hoja de cálculo
        img = openpyxl.drawing.image.Image("graph.png")
        ws.add_image(img, 'I1')

        # Guardar el archivo
        wb.save(file_path)  # Use the file path chosen by the user

        # Convert the Excel file to PDF using Excel's SaveAs function
        pdf_filename = f"{os.path.splitext(file_path)[0]}.pdf"
        excel = win32com.client.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(os.path.abspath(file_path))
        ws_index_list = [ws.Name for ws in workbook.Sheets]  # List of worksheet indices that you want to print
        path_to_pdf = os.path.abspath(pdf_filename)

        # Set the page orientation to landscape and fit to one page
        for ws_index in ws_index_list:
            ws = workbook.Worksheets[ws_index]
            ws.PageSetup.Orientation = 2  # 2 corresponds to landscape orientation
            ws.PageSetup.Zoom = False  # Must be False for FitToPagesWide and FitToPagesTall to work
            ws.PageSetup.FitToPagesWide = 1  # Fit to one page wide
            ws.PageSetup.FitToPagesTall = 1  # Fit to one page tall

        # Select multiple worksheets
        workbook.Sheets([ws_name for ws_name in ws_index_list]).Select()

        workbook.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
        workbook.Close(True)
        excel.Quit()

     

def main():
    root = tk.Tk()
    app = ElectrometroGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()